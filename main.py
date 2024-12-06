import sys
import os
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QPushButton, QLabel, QFileDialog, QVBoxLayout, QHBoxLayout, QScrollArea, QWidget
)
from PyQt5.QtGui import QPixmap, QIcon, QFont, QImage
from PyQt5.QtCore import Qt
from PIL import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage  
from openpyxl.styles import Alignment
from ultralytics import YOLO
import glob
import matplotlib.pyplot as plt
from io import BytesIO

class firstwindow (QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('WatchLeaf')
        self.setWindowIcon(QIcon("iconwindow.png"))

        self.startbut()
        self.showMaximized()
        self.setStyleSheet("""
            QMainWindow {
                background-image: url('BG1.png');
                background-repeat: no-repeat;
                background-position: center;
                background-size: contain;
            }
        """)

    def startbut (self):
        self.pushbutton = QPushButton('start',self)
        self.pushbutton.setGeometry(845,900,300,75)
        self.pushbutton.setStyleSheet('font-size: 25px;')
        self.pushbutton.setStyleSheet("""
        QPushButton {
            border: 0px solid #555;
            border-radius: 20px;  /* Rounded corners */
            background-color: #AF7928;
            color: white;
            font-size: 30px;
            padding: 5px;
        }
        QPushButton:hover {
            background-color: #FFA500;  /* Change color on hover */
        }
        QPushButton:pressed {
            background-color: #E59400;  /* Change color when pressed */
        }
        """)
        self.pushbutton.setFixedSize(300, 75)  

        self.pushbutton.clicked.connect(self.tomainwindow)
        
    def tomainwindow (self):
        self.x = SelectionWindow()
        self.x.show()
        self.close()

class SelectionWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('WatchLeaf - Choose Detection Mode')
        self.setWindowIcon(QIcon("iconwindow.png"))
        self.setStyleSheet("""
            QMainWindow {
                background-image: url('BG1.png');
                background-repeat: no-repeat;
                background-position: center;
                background-size: contain;
            }
            QPushButton {
                border-radius: 15px;
                background-color: #AF7928;
                color: white;
                font-size: 25px;
                padding: 10px;
            }
            QPushButton:hover {
                background-color: #FFA500;
            }
            QPushButton:pressed {
                background-color: #E59400;
            }
            QLabel {
                font-size: 20px;
                color: #333;
            }
        """)

        self.setup_ui()
        self.showMaximized()

    def setup_ui(self):

        self.single_button = QPushButton("Single Image Detection",self)
        self.single_button.clicked.connect(self.open_single_image_detection)
        self.single_button.setGeometry(600,800,300,75)

        self.single_desc = QLabel("Use Single Image Detection when analyzing a single leaf image to detect damage or analyze its condition.",self)
        self.single_desc.setWordWrap(True)
        self.single_desc.setGeometry(600,900,300,100)
        self.single_desc.setAlignment(Qt.AlignCenter)

        self.multiple_button = QPushButton("Multiple Image Detection",self)
        self.multiple_button.clicked.connect(self.open_multiple_image_detection)
        self.multiple_button.setGeometry(1000,800,300,75)

        self.multiple_desc = QLabel("Use Multiple Image Detection when analyzing a batch of images for statistical summaries or comparisons.",self)
        self.multiple_desc.setWordWrap(True)
        self.multiple_desc.setGeometry(1000,900,300,100)
        self.multiple_desc.setAlignment(Qt.AlignCenter)

    def open_single_image_detection(self):
        
        self.single_window = SingleImageWindow()
        self.single_window.show()
        self.close()

    def open_multiple_image_detection(self):
        
        self.multiple_window = MultipleDetectionWindow()
        self.multiple_window.show()
        self.close()

class SingleImageWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('WatchLeaf')
        self.setWindowIcon(QIcon('iconwindow.png'))
        self.model = YOLO('C:\\Users\\HaroldFend\\Campus\\WatchLeaf_Project\\runs\\detect\\train2\\weights\\best.pt')
        self.image_path = None
        self.results = None
        self.result_image_path = None
        self.setStyleSheet("""
            QMainWindow {
                background-image: url('BGsecwin.png');
                background-repeat: no-repeat;
                background-position: center;
                background-size: contain;
            }
        """)
        self.init_ui()
        self.showMaximized()

    def init_ui(self):
        self.upload_button()
        self.save_button()
        self.prev_button()
        self.prelabel()
        self.reslabel()
        self.revlabel()
        self.stalabel()

    def upload_button(self):
        self.uploadbut = QPushButton('Import', self)
        self.uploadbut.setGeometry(1400, 1000, 300, 80)
        self.uploadbut.clicked.connect(self.uploadimag)
        self.uploadbut.setStyleSheet(self.button_style())

    def save_button(self):
        self.savebut = QPushButton('Save To', self)
        self.savebut.setGeometry(1400, 900, 300, 80)
        self.savebut.setEnabled(False)
        self.savebut.clicked.connect(self.savetoexl)
        self.savebut.setStyleSheet(self.button_style())

    def prev_button(self):
        self.prevbut = QPushButton('Previous', self)
        self.prevbut.setGeometry(200, 1000, 300, 80)
        self.prevbut.clicked.connect(self.tofirstwin)
        self.prevbut.setStyleSheet(self.button_style())

    def button_style(self):
        return """
        QPushButton {
            border: 0px solid #555;
            border-radius: 20px;
            background-color: #AF7928;
            color: white;
            font-size: 30px;
            padding: 5px;
        }
        QPushButton:hover {
            background-color: #FFA500;
        }
        QPushButton:pressed {
            background-color: #E59400;
        }
        """

    def prelabel(self):
        self.prelabel = QLabel(self)
        self.prelabel.setGeometry(200, 75, 725, 700)
        self.prelabel.setStyleSheet("""
            QLabel {
                border-radius: 20px;
                background-color: #ffffff;
            }
        """)
        self.prelabel.setAlignment(Qt.AlignCenter)  
        self.preview_text = QLabel('Preview Image', self.prelabel)
        self.preview_text.setAlignment(Qt.AlignHCenter | Qt.AlignTop)
        self.preview_text.setStyleSheet("""
            QLabel {
                color: black;
                font-size: 24px;
                background: white;
                padding: 5px;
            }
        """)
        self.preview_text.setGeometry(10, 0, self.prelabel.width(), 50)


    def reslabel(self):

        self.reslabel = QLabel(self)
        self.reslabel.setGeometry(950, 75, 725, 700)
        self.reslabel.setStyleSheet("""
            QLabel {
                border-radius: 20px;
                background-color: #ffffff;
            }
        """)
        self.reslabel.setAlignment(Qt.AlignCenter)

        self.result_text = QLabel('Result Image', self.reslabel)
        self.result_text.setAlignment(Qt.AlignHCenter | Qt.AlignTop)
        self.result_text.setStyleSheet("""
            QLabel {
                color: black;
                font-size: 24px;
                background-color: white;  /* White background for the text */
                padding: 5px;  /* Add some spacing */
            }
        """)
        self.result_text.setGeometry(0, 0, self.reslabel.width(), 50)


    def revlabel(self):
        self.revlabel = QLabel('Short Review:', self)
        self.revlabel.setGeometry(200, 800, 725, 90)
        self.revlabel.setStyleSheet("""
            QLabel {
                border-radius: 20px;
                background-color: #ffffff;
                color: black;
                font-size: 20px;
                padding: 5px;
            }
        """)
        

    def stalabel(self):
        self.stalabel = QLabel('Status: Choosse a single image to proceed', self)
        self.stalabel.setGeometry(200, 900, 725, 90)
        self.stalabel.setStyleSheet("""
            QLabel {
                border-radius: 20px;
                background-color: #ffffff;
                color: black;
                font-size: 20px;
                padding: 5px;
            }
        """)

    def tofirstwin(self):
        self.y = SelectionWindow()
        self.y.show()
        self.close()

    def uploadimag(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Select Image", "", "Image Files (*.png *.jpg *.jpeg)")
        if file_path:
            pixmap = QPixmap(file_path)
        
            if not pixmap.isNull():
                self.prelabel.setPixmap(pixmap.scaled(self.prelabel.width(), self.prelabel.height(), aspectRatioMode=1))

            self.stalabel.setText("Status: Analyzing image...")
            QApplication.processEvents()

            results = self.model.predict(source=file_path, save=True)

            try:
                
                detected_labels = len(results[0].boxes)
                total_area = 0
                for box in results[0].boxes:
                    x1, y1, x2, y2 = box.xyxy[0]
                    area = (x2 - x1) * (y2 - y1)
                    total_area += area

                orig_width, orig_height = results[0].orig_img.shape[1], results[0].orig_img.shape[0]
                normalized_area = total_area / (orig_width * orig_height)

                if detected_labels <= 5 and normalized_area < 0.05:
                    advice = "Your plant is safe. Keep your plant healthy by maintaining it this way."
                elif 6 <= detected_labels <= 10 or 0.05 <= normalized_area <= 0.15:
                    advice = "Your plant is showing signs of potential issues. Consider monitoring it closely."
                elif 11 <= detected_labels <= 15 or 0.15 < normalized_area <= 0.3:
                    advice = "Your plant is at risk. Take action to address these issues immediately."
                else:
                    advice = "Your plant is in a dangerous situation. Consider applying intensive care or consulting an expert."

                self.detected_labels = detected_labels
                self.normalized_area = normalized_area
                self.advice = advice

                self.revlabel.setText(f"Short Review: Detected {detected_labels} objects with total affected area of {normalized_area:.2%}.\n{advice}")
            except Exception as e:
                self.revlabel.setText(f"Short Review: Error occurred ({str(e)}).")

            result_dir = str(self.model.predictor.save_dir)
            base_name = os.path.basename(file_path).split('.')[0]
            result_files = glob.glob(os.path.join(result_dir, f"{base_name}.*"))
            if result_files:
                self.result_image_path = result_files[0]
                result_pixmap = QPixmap(self.result_image_path)
                if not result_pixmap.isNull():
                    self.reslabel.setPixmap(result_pixmap.scaled(self.reslabel.width(), self.reslabel.height(), aspectRatioMode=1))
                    self.stalabel.setText("Status: Analysis complete. Results displayed.")
                    self.savebut.setEnabled(True)
                else:
                    self.stalabel.setText("Status: Failed to display result image.")
                    self.savebut.setEnabled(False)
            else:
                self.stalabel.setText("Status: Error - Processed result not found.")
                self.savebut.setEnabled(False)
        else:
            self.stalabel.setText("Status: No file selected.")

    def savetoexl(self):
        if not self.result_image_path:
            self.stalabel.setText("Status: No result image to save.")
            return

        save_path, _ = QFileDialog.getSaveFileName(self, "Save Excel File", "", "Excel Files (*.xlsx)")
        if save_path:
            if not save_path.endswith('.xlsx'):
                save_path += '.xlsx'
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Results"

            img = ExcelImage(self.result_image_path)
            img.width, img.height = 400, 400 
            ws.add_image(img, "A1")

            ws["H1"] = "Detection Summary"
            ws["H2"] = "Detected Labels"
            ws["I2"] = self.detected_labels

            ws["H3"] = "Normalized Affected Area"
            ws["I3"] = f"{self.normalized_area:.2%}" 

            ws["H4"] = "Advice"
            ws["I4"] = self.advice

            wb.save(save_path)
            self.stalabel.setText(f"Status: Result saved to {save_path}")
        else:
            self.stalabel.setText("Status: Save operation canceled.")

def pil_to_pixmap(pil_image):
    """Convert a PIL Image to a QPixmap."""
    pil_image = pil_image.convert("RGB")
    data = pil_image.tobytes("raw", "RGB") 
    qimage = QImage(data, pil_image.width, pil_image.height, QImage.Format_RGB888)
    pixmap = QPixmap.fromImage(qimage)
    return pixmap

class MultipleDetectionWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('WatchLeaf')
        self.setWindowIcon(QIcon('iconwindow.png'))
        self.model = YOLO('C:\\Users\\HaroldFend\\Campus\\WatchLeaf_Project\\runs\\detect\\train2\\weights\\best.pt')
        self.image_path = None
        self.results = None
        self.result_image_path = None
        self.setStyleSheet("""
            QMainWindow {
                background-image: url('BGsecwin.png');
                background-repeat: no-repeat;
                background-position: center;
                background-size: contain;
            }
        """)
        self.init_ui()
        self.showMaximized()

    def init_ui(self):
        self.upload_button()
        self.save_button()
        self.prev_button()
        self.prelabel()
        self.stalabel()

    def upload_button(self):
        self.uploadbut = QPushButton('Import', self)
        self.uploadbut.setGeometry(1400, 1000, 300, 80)
        self.uploadbut.clicked.connect(self.uploadimag)
        self.uploadbut.setStyleSheet(self.button_style())

    def save_button(self):
        self.savebut = QPushButton('Save To', self)
        self.savebut.setGeometry(1400, 900, 300, 80)
        self.savebut.setEnabled(False)
        self.savebut.clicked.connect(self.savetoexl)
        self.savebut.setStyleSheet(self.button_style())

    def prev_button(self):
        self.prevbut = QPushButton('Previous', self)
        self.prevbut.setGeometry(200, 1000, 300, 80)
        self.prevbut.clicked.connect(self.tofirstwin)
        self.prevbut.setStyleSheet(self.button_style())

    def button_style(self):
        return """
        QPushButton {
            border: 0px solid #555;
            border-radius: 20px;
            background-color: #AF7928;
            color: white;
            font-size: 30px;
            padding: 5px;
        }
        QPushButton:hover {
            background-color: #FFA500;
        }
        QPushButton:pressed {
            background-color: #E59400;
        }
        """

    def prelabel(self):
        self.prelabel = QLabel(self)
        self.prelabel.setGeometry(200, 75, 1475, 700)
        self.prelabel.setStyleSheet("""
            QLabel {
                background-color: #ffffff;
            }
        """)
        self.prelabel.setAlignment(Qt.AlignCenter)
        
        self.preview_text = QLabel('Result', self.prelabel)
        self.preview_text.setAlignment(Qt.AlignHCenter | Qt.AlignTop)
        self.preview_text.setStyleSheet("""
            QLabel {
                color: black;
                font-size: 24px;
                background: white;
                padding: 5px;
            }
        """)
        self.preview_text.setGeometry(10, 0, self.prelabel.width(), 50)

    def stalabel(self):
        self.stalabel = QLabel('Status: Choosse a multiple image to proceed', self)
        self.stalabel.setGeometry(200, 800, 725, 90)
        self.stalabel.setStyleSheet("""
            QLabel {
                border-radius: 20px;
                background-color: #ffffff;
                color: black;
                font-size: 20px;
                padding: 5px;
            }
        """)

    def tofirstwin(self):
        self.window = SelectionWindow()
        self.window.show()
        self.close()

    def uploadimag(self):
        file_paths, _ = QFileDialog.getOpenFileNames(self, "Select Images", "", "Image Files (*.png *.jpg *.jpeg)")
        if not file_paths:
            self.stalabel.setText("Status: No files selected.")
            return
        
        self.result_summary = []
        self.result_images = []

        self.stalabel.setText("Status: Processing images...")
        QApplication.processEvents()

        for file_path in file_paths:
            pixmap = QPixmap(file_path)
            if not pixmap.isNull():
                self.prelabel.setPixmap(pixmap.scaled(self.prelabel.width(), self.prelabel.height(), aspectRatioMode=1))

            try:
                results = self.model.predict(source=file_path, save=True)
                detected_labels = len(results[0].boxes)
                total_area = sum(
                    (box.xyxy[0][2] - box.xyxy[0][0]) * (box.xyxy[0][3] - box.xyxy[0][1])
                    for box in results[0].boxes
                )
                orig_width, orig_height = results[0].orig_img.shape[1], results[0].orig_img.shape[0]
                normalized_area = total_area / (orig_width * orig_height)

                if detected_labels <= 5 and normalized_area < 0.05:
                    advice = "Your plant is safe. Keep your plant healthy by maintaining it this way."
                elif 6 <= detected_labels <= 10 or 0.05 <= normalized_area <= 0.15:
                    advice = "Your plant is showing signs of potential issues. Consider monitoring it closely."
                elif 11 <= detected_labels <= 15 or 0.15 < normalized_area <= 0.3:
                    advice = "Your plant is at risk. Take action to address these issues immediately."
                else:
                    advice = "Your plant is in a dangerous situation. Consider applying intensive care or consulting an expert."

                self.result_summary.append((file_path, detected_labels, normalized_area, advice))

                result_dir = str(self.model.predictor.save_dir)
                base_name = os.path.basename(file_path).split('.')[0]
                result_files = glob.glob(os.path.join(result_dir, f"{base_name}.*"))
                if result_files:
                    self.result_images.append(result_files[0])
                else:
                    self.result_images.append(None)
            except Exception as e:
                self.result_summary.append((file_path, "Error", "Error", str(e)))
                self.result_images.append(None)

        self.display_results()
        self.stalabel.setText("Status: Processing complete.")
        self.savebut.setEnabled(True)


    def generate_chart(self):
        """Generate a chart based on detection results."""
        if not hasattr(self, 'result_summary') or not self.result_summary:
            return None

        file_names = [os.path.basename(result[0]) for result in self.result_summary]
        detected_labels = [result[1] for result in self.result_summary]
        normalized_areas = [result[2] for result in self.result_summary]

        plt.figure(figsize=(10, 6))
        plt.bar(file_names, detected_labels, color='blue', alpha=0.7, label='Detected Labels')
        plt.plot(file_names, normalized_areas, color='red', marker='o', label='Normalized Area')

        plt.xlabel('Images')
        plt.xticks(rotation=45, ha='right')
        plt.ylabel('Count / Area')
        plt.title('Detection Summary')
        plt.legend()
        plt.tight_layout()

        buffer = BytesIO()
        plt.savefig(buffer, format='png')
        buffer.seek(0)
        plt.close()

        chart_image = Image.open(buffer)
        return chart_image

    def display_results(self):
        if hasattr(self, "results_container"):
            self.results_container.deleteLater() 

        self.results_container = QScrollArea(self)
        self.results_container.setGeometry(200, 75, 1475, 700)
        self.results_container.setStyleSheet("background-color: white; border-radius: 20px;")

        results_widget = QWidget()
        results_layout = QVBoxLayout()

        image_width, image_height = 500, 500 
        font = QFont()
        font.setPointSize(12)


        chart_image = self.generate_chart()
        
        if chart_image:
            chart_label = QLabel()
            chart_pixmap = pil_to_pixmap(chart_image) 
            chart_label.setPixmap(chart_pixmap.scaled(1000, 500, aspectRatioMode=1))
            chart_label.setAlignment(Qt.AlignCenter)
            results_layout.addWidget(chart_label)

        for idx, (file_path, detected_labels, normalized_area, advice) in enumerate(self.result_summary):

            result_container = QWidget()
            result_layout = QHBoxLayout()
            result_container.setLayout(result_layout)

            preview_pixmap = QPixmap(file_path)
            preview_label = QLabel()
            preview_label.setPixmap(preview_pixmap.scaled(image_width, image_height, aspectRatioMode=1))
            preview_label.setAlignment(Qt.AlignCenter)
            result_layout.addWidget(preview_label)

            result_image_label = QLabel()
            if self.result_images[idx]:
                result_pixmap = QPixmap(self.result_images[idx])
                result_image_label.setPixmap(result_pixmap.scaled(image_width, image_height, aspectRatioMode=1))
            else:
                result_image_label.setText("No result image available")
            result_image_label.setAlignment(Qt.AlignCenter)
            result_layout.addWidget(result_image_label)

            result_label = QLabel()
            result_label.setFont(font)
            if detected_labels == "Error":
                result_label.setText(f"Error processing {os.path.basename(file_path)}: {advice}")
            else:
                result_label.setText(
                    f"Image: {os.path.basename(file_path)}\n"
                    f"Detected Labels: {detected_labels}\n"
                    f"Normalized Area: {normalized_area:.2%}\n"
                    f"Advice: {advice}"
                )
            result_label.setAlignment(Qt.AlignLeft)
            result_layout.addWidget(result_label)


            result_layout.setSpacing(20)

            results_layout.addWidget(result_container)
            results_layout.addSpacing(30) 


        results_widget.setLayout(results_layout)
        self.results_container.setWidget(results_widget)
        self.results_container.show()

    def savetoexl(self):
        if not hasattr(self, 'result_summary') or not self.result_summary:
            self.stalabel.setText("Status: No results to save.")
            return

        save_path, _ = QFileDialog.getSaveFileName(self, "Save Excel File", "", "Excel Files (*.xlsx)")
        if not save_path:
            self.stalabel.setText("Status: Save operation canceled.")
            return

        if not save_path.endswith('.xlsx'):
            save_path += '.xlsx'

        wb = Workbook()
        ws = wb.active
        ws.title = "Results"


        ws.append(["Image", "Image", "Image", "Image", "Summary"])


        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 50 

        start_row = 2

        for idx, (file_path, detected_labels, normalized_area, advice) in enumerate(self.result_summary):
            if detected_labels == "Error":
                continue

            result_dir = str(self.model.predictor.save_dir)
            base_name = os.path.basename(file_path).split('.')[0]
            result_files = glob.glob(os.path.join(result_dir, f"{base_name}.*"))
            if not result_files:
                continue

            result_image_path = result_files[0]

            try:
                img = ExcelImage(result_image_path)
                img.width, img.height = 400, 400
                ws.add_image(img, f"A{start_row}")

                ws[f"E{start_row}"] = f"Image: {os.path.basename(file_path)}"
                ws[f"E{start_row + 1}"] = f"Detected Labels: {detected_labels}"
                ws[f"E{start_row + 2}"] = f"Normalized Affected Area: {normalized_area:.2%}"
                ws[f"E{start_row + 3}"] = f"Advice: {advice}"

                ws[f"E{start_row + 3}"].alignment = Alignment(wrap_text=True)

            except Exception as e:
                print(f"Error inserting image {file_path}: {e}")
                continue

            start_row += 17

        chart_image = self.generate_chart()
        if chart_image:
            chart_path = os.path.join(os.path.dirname(save_path), "chart.png")
            chart_image.save(chart_path)

            try:
                img = ExcelImage(chart_path)
                img.width, img.height = 800, 400
                ws.add_image(img, f"A{start_row}")
                start_row += 20 
            except Exception as e:
                print(f"Error adding chart to Excel: {e}")

        try:
            wb.save(save_path)
            self.stalabel.setText(f"Status: Results saved to {save_path}")
        except Exception as e:
            print(f"Error saving the Excel file: {e}")
            self.stalabel.setText("Status: Error saving the Excel file.")

app = QApplication(sys.argv)
window = firstwindow()
window.show()
app.exec_()